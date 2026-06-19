"""
Validador Pp - Partida Específica
Aplicación para validar combinaciones de Programa presupuestario y Partida
según el catálogo oficial de SADER.
"""

import streamlit as st
import pandas as pd
import pickle
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE PÁGINA
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Validador Pp-Partida | SADER",
    page_icon="✓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ══════════════════════════════════════════════════════════════════════════════
# CONTRASEÑA
# ══════════════════════════════════════════════════════════════════════════════

PASSWORD = "#sader 2026"

def check_password():
    """Muestra pantalla de acceso y devuelve True si la contraseña es correcta."""
    if st.session_state.get("autenticado"):
        return True

    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@400;600;700&display=swap');
        .login-box {
            max-width: 420px;
            margin: 6rem auto;
            background: white;
            border-radius: 14px;
            padding: 2.5rem 2rem;
            box-shadow: 0 6px 24px rgba(107,29,61,0.15);
            text-align: center;
        }
        .login-logo {
            background: linear-gradient(135deg, #6B1D3D 0%, #8B2D4D 100%);
            color: white;
            border-radius: 10px;
            padding: 1rem 1.5rem;
            margin-bottom: 1.5rem;
        }
        .login-logo h2 { margin: 0; font-size: 1.4rem; }
        .login-logo p  { margin: 0.3rem 0 0 0; opacity: 0.85; font-size: 0.9rem; }
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("""
        <div class="login-box">
            <div class="login-logo">
                <h2>✓ Validador Pp-Partida</h2>
                <p>SADER · Secretaría de Agricultura y Desarrollo Rural</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        pwd = st.text_input("Contraseña de acceso", type="password", placeholder="Ingresa la contraseña")
        if st.button("Entrar", type="primary", use_container_width=True):
            if pwd == PASSWORD:
                st.session_state["autenticado"] = True
                st.rerun()
            else:
                st.error("Contraseña incorrecta")
    return False


# ══════════════════════════════════════════════════════════════════════════════
# PERSISTENCIA DE CATÁLOGOS
# ══════════════════════════════════════════════════════════════════════════════

PERSIST_DIR = "data_persistente"
os.makedirs(PERSIST_DIR, exist_ok=True)

CATALOGOS_PICKLE = os.path.join(PERSIST_DIR, "catalogos_validador.pkl")

def guardar_catalogos(catalogos: dict):
    """Guarda los catálogos en disco para persistencia entre sesiones."""
    with open(CATALOGOS_PICKLE, "wb") as f:
        pickle.dump(catalogos, f)

def cargar_catalogos_disco() -> dict:
    """Carga los catálogos desde disco si existen."""
    if os.path.exists(CATALOGOS_PICKLE):
        try:
            with open(CATALOGOS_PICKLE, "rb") as f:
                return pickle.load(f)
        except Exception:
            return {}
    return {}


# ══════════════════════════════════════════════════════════════════════════════
# ESTILOS CSS PERSONALIZADOS
# ══════════════════════════════════════════════════════════════════════════════

def inyectar_css():
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@400;600;700&display=swap');
        :root {
            --guinda: #6B1D3D;
            --guinda-claro: #8B2D4D;
            --crema: #F5F0E6;
            --verde-ok: #2E7D32;
            --rojo-error: #C62828;
        }
        .main-header {
            background: linear-gradient(135deg, var(--guinda) 0%, var(--guinda-claro) 100%);
            color: white;
            padding: 1.5rem 2rem;
            border-radius: 10px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 15px rgba(107, 29, 61, 0.3);
        }
        .main-header h1 { margin: 0; font-size: 1.8rem; font-weight: 700; }
        .main-header p  { margin: 0.5rem 0 0 0; opacity: 0.9; font-size: 1rem; }
        .stat-card {
            background: white; border-radius: 10px; padding: 1.2rem;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            border-left: 4px solid var(--guinda); margin-bottom: 1rem;
        }
        .stat-card.success { border-left-color: var(--verde-ok); }
        .stat-card.error   { border-left-color: var(--rojo-error); }
        .stat-number { font-size: 2rem; font-weight: 700; color: var(--guinda); line-height: 1; }
        .stat-label  { color: #666; font-size: 0.9rem; margin-top: 0.3rem; }
        .result-valid {
            background: #E8F5E9; border: 1px solid #A5D6A7;
            border-radius: 8px; padding: 1rem; margin: 0.5rem 0;
        }
        .result-invalid {
            background: #FFEBEE; border: 1px solid #EF9A9A;
            border-radius: 8px; padding: 1rem; margin: 0.5rem 0;
        }
        .badge-persistente {
            display: inline-block;
            background: #E8F5E9; color: #2E7D32;
            border: 1px solid #A5D6A7; border-radius: 20px;
            padding: 0.2rem 0.7rem; font-size: 0.78rem; font-weight: 600;
        }
        .stButton > button {
            background: var(--guinda); color: white; border: none;
            border-radius: 6px; padding: 0.5rem 1.5rem;
            font-weight: 600; transition: all 0.3s ease;
        }
        .stButton > button:hover {
            background: var(--guinda-claro);
            box-shadow: 0 4px 12px rgba(107, 29, 61, 0.3);
        }
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE PROCESAMIENTO
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data
def cargar_catalogo(archivo_bytes: bytes) -> dict:
    """Carga y procesa el catálogo de Pp-Partidas desde bytes."""
    df = pd.read_excel(BytesIO(archivo_bytes), header=None, dtype=str)
    df = df.iloc[1:].reset_index(drop=True)
    partidas_por_pp = {}
    for _, row in df.iterrows():
        mod     = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
        prog    = str(row.iloc[4]).strip().zfill(3) if pd.notna(row.iloc[4]) else ''
        partida = str(row.iloc[6]).strip().zfill(5) if pd.notna(row.iloc[6]) else ''
        if mod and prog and partida and partida not in ('nan', '00nan'):
            pp = f"{mod}{prog}"
            partidas_por_pp.setdefault(pp, set()).add(partida)
    return partidas_por_pp


def procesar_archivo_validacion(archivo, partidas_por_pp):
    """Procesa archivo de claves a validar."""
    df_raw = pd.read_excel(archivo, header=None, dtype=str)
    datos = []
    fila_datos = None
    for i in range(min(15, len(df_raw))):
        val0 = str(df_raw.iloc[i, 0]).strip() if pd.notna(df_raw.iloc[i, 0]) else ''
        val1 = str(df_raw.iloc[i, 1]).strip() if df_raw.shape[1] > 1 and pd.notna(df_raw.iloc[i, 1]) else ''
        if val0.isdigit() and len(val0) <= 2 and int(val0) > 0:
            fila_datos = i; break
        if val1.isdigit() and len(val1) <= 2 and int(val1) > 0:
            fila_datos = i; break

    if fila_datos is not None:
        df_datos = df_raw.iloc[fila_datos:].reset_index(drop=True)
        for _, row in df_datos.iterrows():
            pp_val      = str(row.iloc[9]).strip().upper()  if len(row) > 9  and pd.notna(row.iloc[9])  else ''
            partida_val = str(row.iloc[10]).strip().zfill(5) if len(row) > 10 and pd.notna(row.iloc[10]) else ''
            if pp_val and pp_val not in ('nan','NAN') and partida_val and partida_val != '0000n':
                datos.append({'PP': pp_val, 'PARTIDA': partida_val})
    else:
        df_cols = pd.read_excel(archivo, dtype=str)
        col_pp = col_partida = None
        for c in df_cols.columns:
            cu = str(c).upper()
            if 'PP' in cu or 'PROGRAMA' in cu: col_pp = c
            if 'PARTIDA' in cu or 'OBJETO' in cu: col_partida = c
        if col_pp and col_partida:
            for _, row in df_cols.iterrows():
                pp_val      = str(row[col_pp]).strip().upper()   if pd.notna(row[col_pp])      else ''
                partida_val = str(row[col_partida]).strip().zfill(5) if pd.notna(row[col_partida]) else ''
                if pp_val and pp_val not in ('nan','NAN'):
                    datos.append({'PP': pp_val, 'PARTIDA': partida_val})
    return datos


def validar_registros(datos, partidas_por_pp):
    resultados = []
    for d in datos:
        pp, partida = d['PP'], d['PARTIDA']
        if pp not in partidas_por_pp:
            resultados.append({'PP': pp, 'PARTIDA': partida, 'VÁLIDO': 'NO', 'MOTIVO': f'Pp {pp} no existe en catálogo'})
        elif partida in partidas_por_pp[pp]:
            resultados.append({'PP': pp, 'PARTIDA': partida, 'VÁLIDO': 'SÍ', 'MOTIVO': ''})
        else:
            resultados.append({'PP': pp, 'PARTIDA': partida, 'VÁLIDO': 'NO', 'MOTIVO': 'Partida no autorizada para este Pp'})
    return resultados


def generar_excel_resultados(resultados):
    wb = Workbook(); ws = wb.active; ws.title = "Validación"
    header_fill = PatternFill(start_color='6B1D3D', end_color='6B1D3D', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    si_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    no_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border  = Border(left=Side(style='thin',color='CCCCCC'), right=Side(style='thin',color='CCCCCC'),
                     top=Side(style='thin',color='CCCCCC'),  bottom=Side(style='thin',color='CCCCCC'))
    center  = Alignment(horizontal='center', vertical='center')
    for col, h in enumerate(['PP','PARTIDA','VÁLIDO','MOTIVO'], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.border, c.alignment = header_fill, header_font, border, center
    for i, r in enumerate(resultados, 2):
        ws.cell(row=i, column=1, value=r['PP']).border = border
        ws.cell(row=i, column=2, value=r['PARTIDA']).border = border
        cv = ws.cell(row=i, column=3, value=r['VÁLIDO'])
        cv.border, cv.alignment = border, center
        cv.fill = si_fill if r['VÁLIDO'] == 'SÍ' else no_fill
        ws.cell(row=i, column=4, value=r['MOTIVO']).border = border
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    output = BytesIO(); wb.save(output); output.seek(0)
    return output


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR - CARGA / GESTIÓN DE CATÁLOGO
# ══════════════════════════════════════════════════════════════════════════════

def sidebar_catalogo():
    """Renderiza el sidebar y devuelve el dict partidas_por_pp activo."""
    # Inicializar catalogos en session_state desde disco si aún no está
    if "catalogos" not in st.session_state:
        st.session_state["catalogos"] = cargar_catalogos_disco()

    catalogos = st.session_state["catalogos"]
    partidas_por_pp = catalogos.get("pp_partida", {})

    with st.sidebar:
        st.markdown("###  Catálogo Base")

        if partidas_por_pp:
            meta = catalogos.get("pp_partida_meta", {})
            st.markdown(
                f'<span class="badge-persistente">✓ Catálogo guardado</span>',
                unsafe_allow_html=True
            )
            if meta.get("nombre"):
                st.caption(f"Archivo: **{meta['nombre']}**")
            if meta.get("fecha"):
                st.caption(f"Cargado: {meta['fecha']}")
            st.metric("Programas (Pp)", len(partidas_por_pp))
            st.metric("Total partidas", sum(len(v) for v in partidas_por_pp.values()))

            st.markdown("---")
            with st.expander(" Actualizar catálogo"):
                st.caption("Sube un nuevo archivo para reemplazar el catálogo guardado")
                nuevo = st.file_uploader(
                    "Nuevo catálogo Pp-Partida",
                    type=['xlsx','xls'],
                    key="catalogo_nuevo",
                    label_visibility="collapsed"
                )
                if nuevo:
                    with st.spinner("Procesando..."):
                        nuevo_datos = cargar_catalogo(nuevo.read())
                    from datetime import datetime
                    catalogos["pp_partida"] = nuevo_datos
                    catalogos["pp_partida_meta"] = {
                        "nombre": nuevo.name,
                        "fecha": datetime.now().strftime("%d/%m/%Y %H:%M")
                    }
                    guardar_catalogos(catalogos)
                    st.session_state["catalogos"] = catalogos
                    st.success("✓ Catálogo actualizado y guardado")
                    st.rerun()
        else:
            st.caption("Sube el archivo `Pp_-_Partida_Especifica_2026.xlsx`")
            catalogo_file = st.file_uploader(
                "Catálogo Pp-Partida",
                type=['xlsx','xls'],
                key="catalogo_inicial",
                label_visibility="collapsed"
            )
            if catalogo_file:
                with st.spinner("Cargando catálogo..."):
                    partidas_por_pp = cargar_catalogo(catalogo_file.read())
                from datetime import datetime
                catalogos["pp_partida"] = partidas_por_pp
                catalogos["pp_partida_meta"] = {
                    "nombre": catalogo_file.name,
                    "fecha": datetime.now().strftime("%d/%m/%Y %H:%M")
                }
                guardar_catalogos(catalogos)
                st.session_state["catalogos"] = catalogos
                st.success("✓ Catálogo guardado")
                st.rerun()

        st.markdown("---")
        st.markdown("### Pps disponibles")
        if partidas_por_pp:
            pps_lista = sorted(partidas_por_pp.keys())
            st.text_area("", value=", ".join(pps_lista), height=150, label_visibility="collapsed")

        # Cerrar sesión
        st.markdown("---")
        if st.button(" Cerrar sesión", use_container_width=True):
            st.session_state["autenticado"] = False
            st.rerun()

    return partidas_por_pp


# ══════════════════════════════════════════════════════════════════════════════
# INTERFAZ PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

if not check_password():
    st.stop()

inyectar_css()

st.markdown("""
<div class="main-header">
    <h1>✓ Validador Pp - Partida Específica</h1>
    <p>Verifica combinaciones de Programa Presupuestario y Partida según el catálogo oficial</p>
</div>
""", unsafe_allow_html=True)

partidas_por_pp = sidebar_catalogo()

if not partidas_por_pp:
    st.info(" **Primero sube el catálogo** `Pp_-_Partida_Especifica_2026.xlsx` en la barra lateral para comenzar.")
    st.stop()

tab1, tab2, tab3 = st.tabs([
    " Consulta Individual",
    " Validación Masiva",
    " Explorar Catálogo"
])

# ══ TAB 1: CONSULTA INDIVIDUAL ════════════════════════════════════════════════
with tab1:
    st.markdown("#### Validar una combinación Pp + Partida")
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        pp_input = st.text_input("Programa Presupuestario (Pp)", placeholder="Ej: S263, K017, E009", max_chars=10).upper().strip()
    with col2:
        partida_input = st.text_input("Partida específica", placeholder="Ej: 33104, 52301", max_chars=5, key="partida_input_value").strip()
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        validar_btn = st.button("Validar", type="primary", use_container_width=True)

    if validar_btn and pp_input:
        partida_check = partida_input.zfill(5) if partida_input else ""
        if pp_input not in partidas_por_pp:
            st.markdown(f'<div class="result-invalid"><strong>❌ Pp no encontrado</strong><br>El programa <code>{pp_input}</code> no existe en el catálogo.</div>', unsafe_allow_html=True)
            similares = [p for p in partidas_por_pp.keys() if pp_input[0] in p][:5]
            if similares:
                st.caption(f"¿Quisiste decir?: {', '.join(similares)}")
        elif not partida_check or partida_check == "00000":
            partidas = sorted(partidas_por_pp[pp_input])
            st.success(f"✓ **Pp {pp_input}** tiene **{len(partidas)}** partidas válidas")
            capitulos = {}
            for p in partidas:
                capitulos.setdefault(p[0], []).append(p)
            for cap in sorted(capitulos.keys()):
                with st.expander(f"Capítulo {cap}000 ({len(capitulos[cap])} partidas)"):
                    st.markdown(" ".join([f"`{p}`" for p in capitulos[cap]]))
        elif partida_check in partidas_por_pp[pp_input]:
            st.markdown(f'<div class="result-valid"><strong> VÁLIDO</strong><br>La partida <code>{partida_check}</code> está autorizada para el Pp <code>{pp_input}</code></div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="result-invalid"><strong> NO VÁLIDO</strong><br>La partida <code>{partida_check}</code> <strong>no</strong> está autorizada para el Pp <code>{pp_input}</code></div>', unsafe_allow_html=True)
            cap = partida_check[0]
            similares = sorted([p for p in partidas_por_pp[pp_input] if p[0] == cap])
            if similares:
                st.caption(f"Partidas válidas del capítulo {cap}000:")
                st.code(", ".join(similares[:20]))

# ══ TAB 2: VALIDACIÓN MASIVA ══════════════════════════════════════════════════
with tab2:
    st.markdown("#### Validar múltiples registros desde archivo")
    st.caption("Soporta formato PIPP o archivos con columnas Pp/Partida")
    archivo_validar = st.file_uploader("Archivo con claves a validar", type=['xlsx','xls'], key="validar")
    if archivo_validar:
        with st.spinner("Procesando archivo..."):
            datos = procesar_archivo_validacion(archivo_validar, partidas_por_pp)
        if not datos:
            st.error("No se encontraron registros válidos en el archivo")
        else:
            st.info(f"**{len(datos)}** registros encontrados")
            if st.button("✔ Validar registros", type="primary"):
                resultados = validar_registros(datos, partidas_por_pp)
                validos   = sum(1 for r in resultados if r['VÁLIDO'] == 'SÍ')
                invalidos = len(resultados) - validos
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.markdown(f'<div class="stat-card"><div class="stat-number">{len(resultados)}</div><div class="stat-label">Total registros</div></div>', unsafe_allow_html=True)
                with col2:
                    st.markdown(f'<div class="stat-card success"><div class="stat-number" style="color:#2E7D32">{validos}</div><div class="stat-label">Válidos ✓</div></div>', unsafe_allow_html=True)
                with col3:
                    st.markdown(f'<div class="stat-card error"><div class="stat-number" style="color:#C62828">{invalidos}</div><div class="stat-label">Con errores ✗</div></div>', unsafe_allow_html=True)
                st.markdown("---")
                st.markdown("##### Detalle de validación")
                df_res = pd.DataFrame(resultados)
                def highlight_valid(row):
                    return ['background-color: #E8F5E9' if row['VÁLIDO']=='SÍ' else 'background-color: #FFEBEE'] * len(row)
                st.dataframe(df_res.style.apply(highlight_valid, axis=1), use_container_width=True, height=400)
                excel_output = generar_excel_resultados(resultados)
                st.download_button(
                    label="⬇ Descargar resultados (.xlsx)",
                    data=excel_output,
                    file_name="Validacion_Pp_Partida.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# ══ TAB 3: EXPLORAR CATÁLOGO ══════════════════════════════════════════════════
with tab3:
    st.markdown("#### Explorar catálogo completo")
    pp_seleccionado = st.selectbox(
        "Selecciona un Programa Presupuestario",
        options=[""] + sorted(partidas_por_pp.keys()),
        format_func=lambda x: f"{x} ({len(partidas_por_pp.get(x,[]))} partidas)" if x else "-- Seleccionar --"
    )
    if pp_seleccionado:
        partidas = sorted(partidas_por_pp[pp_seleccionado])
        st.success(f"**{pp_seleccionado}** tiene **{len(partidas)}** partidas autorizadas")
        capitulos = {}
        for p in partidas:
            capitulos.setdefault(p[0], []).append(p)
        nombres_cap = {'1':'Servicios Personales','2':'Materiales y Suministros','3':'Servicios Generales',
                       '4':'Transferencias','5':'Bienes Muebles','6':'Inversión Pública',
                       '7':'Inversiones Financieras','8':'Participaciones','9':'Deuda Pública'}
        for cap in sorted(capitulos.keys()):
            with st.expander(f"**Capítulo {cap}000** - {nombres_cap.get(cap,'')} ({len(capitulos[cap])} partidas)", expanded=True):
                cols = st.columns(6)
                for i, partida in enumerate(capitulos[cap]):
                    cols[i % 6].code(partida)

# ══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("---")
st.caption("Validador Pp-Partida | SADER - Secretaría de Agricultura y Desarrollo Rural")
